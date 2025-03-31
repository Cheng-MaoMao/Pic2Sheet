import os
from openai import OpenAI
import openpyxl
from datetime import datetime
import json

class ImageAnalyzer:
    def __init__(self, provider="阿里", api_key=None):
        if not api_key:
            raise ValueError("API Key 不能为空，请在配置中设置有效的 API Key")
            
        config_path = "config.json"
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
        except Exception:
            # 使用默认配置
            self.config = {
                "阿里": {
                    "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
                    "model": "qwen-vl-max-latest"
                },
                "火山引擎": {
                    "base_url": "https://ark.cn-beijing.volces.com/api/v3",
                    "model": "doubao-1-5-vision-pro-32k-250115"
                }
            }
        
        if provider not in self.config:
            raise ValueError("不支持的提供商")
            
        provider_config = self.config[provider]
        self.provider = provider
        self.client = OpenAI(
            api_key=api_key,
            base_url=provider_config["base_url"]
        )
        self.model = provider_config["model"]
        
    def analyze_image(self, image_path):
        prompt = """请分析图片内容：
1. 分析图片，如果图片包含表格，请严格按照图片中的表格格式返回数据(保留空的内容)，以JSON数组形式，例如：
[["表头1", "表头2"], ["数据1", "数据2"]]

2. 分析图片，如果图片不包含表格，请分析内容并将相关信息整理成表格形式，以JSON数组返回，例如：
[["类别", "内容"], ["姓名", "张三"], ["年龄", "25"]]

请确保返回格式为标准JSON数组。"""

        if image_path.startswith(('http://', 'https://')):
            image_url = image_path
        else:
            import base64
            with open(image_path, "rb") as image_file:
                image_data = base64.b64encode(image_file.read()).decode('utf-8')
                image_url = f"data:image/jpeg;base64,{image_data}"
        
        try:
            completion = self.client.chat.completions.create(
                model=self.model,  # 使用配置文件中的模型名称
                temperature = 0.7,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": image_url}}
                    ]
                }]
            )
            
            response_data = json.loads(completion.model_dump_json())
            content = response_data['choices'][0]['message']['content']
            
            # 打印原始返回内容，用于调试
            print("原始返回内容:", content)
            print("提供商:", self.provider)
            print("使用的模型:", self.model)
            
            # 尝试清理和格式化内容
            content = content.strip()
            if content.startswith('```') and content.endswith('```'):
                content = content[3:-3].strip()
            if content.startswith('json'):
                content = content[4:].strip()
                
            return content
            
        except Exception as e:
            print(f"分析图片时发生错误: {str(e)}")
            return None

    def analyze_images_batch(self, image_paths, batch_size=5, max_retries=3):
        """批量分析多张关联图片，使用分批处理机制"""
        all_results = []
        total_images = len(image_paths)
        
        # 分批处理图片
        for i in range(0, total_images, batch_size):
            batch_paths = image_paths[i:min(i+batch_size, total_images)]
            
            # 添加重试机制
            for retry in range(max_retries):
                try:
                    print(f"处理第{i+1}到第{min(i+batch_size, total_images)}批图片 (尝试 {retry + 1}/{max_retries})")
                    
                    prompt = f"""请分析以下第{i+1}到第{min(i+batch_size, total_images)}张图片内容：

图片编号：
"""
                    # 添加当前批次的图片编号
                    for idx, path in enumerate(batch_paths, i+1):
                        prompt += f"[{idx}] {os.path.basename(path)}\n"
                    
                    prompt += """
请按照以下要求处理：
1. 提取每张图片中的信息并保持完整性
2. 使用统一的表头格式
3. 每一行代表一条完整记录
4. 保持数据的完整性，包括空值
5. 仅返回JSON格式的数据数组

请直接返回JSON数组，格式如下：
[
    ["表头1", "表头2", "表头3", ...],
    ["数据1", "数据2", "数据3", ...],
    ["数据1", "数据2", "数据3", ...]
]"""

                    all_image_data = []
                    # 处理当前批次的图片
                    for image_path in batch_paths:
                        if image_path.startswith(('http://', 'https://')):
                            image_url = image_path
                        else:
                            import base64
                            with open(image_path, "rb") as image_file:
                                image_data = base64.b64encode(image_file.read()).decode('utf-8')
                                image_url = f"data:image/jpeg;base64,{image_data}"
                        all_image_data.append({"type": "image_url", "image_url": {"url": image_url}})

                    completion = self.client.chat.completions.create(
                        model=self.model,
                        temperature=0.7,
                        messages=[{
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                *all_image_data
                            ]
                        }],
                        max_tokens=4096,
                        timeout=120  # 增加超时时间到120秒
                    )
                    
                    response_data = json.loads(completion.model_dump_json())
                    content = response_data['choices'][0]['message']['content']
                    
                    # 清理和格式化内容
                    content = content.strip()
                    if content.startswith('```') and content.endswith('```'):
                        content = content[3:-3].strip()
                    if content.startswith('json'):
                        content = content[4:].strip()
                    
                    # 验证JSON格式
                    batch_result = json.loads(content)
                    if not isinstance(batch_result, list):
                        raise ValueError("返回的数据格式不是数组")
                    
                    if batch_result and isinstance(batch_result, list):
                        if not all_results:
                            # 第一批，保留表头
                            all_results.extend(batch_result)
                        else:
                            # 后续批次，只添加数据行
                            all_results.extend(batch_result[1:])
                    
                    print(f"已处理 {min(i+batch_size, total_images)}/{total_images} 张图片")
                    # 处理成功，跳出重试循环
                    break
                    
                except Exception as e:
                    print(f"处理第{i+1}到第{min(i+batch_size, total_images)}批图片时发生错误: {str(e)}")
                    if retry < max_retries - 1:
                        print("准备重试...")
                        import time
                        time.sleep(2)  # 等待2秒后重试
                        continue
                    else:
                        print("已达到最大重试次数，跳过该批次")
                        return None  # 返回None表示处理失败
        
        # 确保有处理结果才返回
        if all_results:
            return json.dumps(all_results, ensure_ascii=False, indent=2)
        return None

    def get_format_suggestions(self, data):
        """询问大模型关于数据格式化的建议"""
        try:
            # 将数据转换为字符串形式
            data_str = json.dumps(data, ensure_ascii=False, indent=2)
            
            prompt = f"""请分析以下数据表格，并给出Excel格式建议：
数据内容：
{data_str}

请严格按照以下JSON格式返回建议：
{{
    "columns": [
        {{
            "header": "列名",
            "width": 15,
            "alignment": "left",
            "wrap_text": false,
            "format": "常规",
            "color": "#E6E6E6"
        }}
    ]
}}

注意：
1. 必须返回合法的JSON格式
2. 所有属性值必须符合规范
3. 颜色值使用16进制格式
4. width必须是数字
5. alignment只能是 "left"、"center" 或 "right"
6. wrap_text必须是true或false
"""

            completion = self.client.chat.completions.create(
                model=self.model,
                temperature=0.3,  # 降低温度以获得更稳定的输出
                messages=[{
                    "role": "user",
                    "content": prompt
                }]
            )
            
            response_data = json.loads(completion.model_dump_json())
            content = response_data['choices'][0]['message']['content']
            
            # 打印原始返回内容以便调试
            print("格式建议原始返回:", content)
            
            # 清理JSON字符串
            content = content.strip()
            # 提取JSON部分
            import re
            json_match = re.search(r'\{[\s\S]*\}', content)
            if json_match:
                json_str = json_match.group()
                # 尝试解析和验证JSON
                try:
                    format_data = json.loads(json_str)
                    # 验证数据结构
                    if not isinstance(format_data, dict) or 'columns' not in format_data:
                        raise ValueError("格式建议数据结构无效")
                    return format_data
                except json.JSONDecodeError as e:
                    print(f"JSON解析错误: {str(e)}")
                    print("问题JSON:", json_str)
                    # 返回默认格式
                    return self._get_default_format_suggestions(data)
            else:
                print("未找到有效的JSON格式")
                return self._get_default_format_suggestions(data)
            
        except Exception as e:
            print(f"获取格式建议时发生错误: {str(e)}")
            return self._get_default_format_suggestions(data)

    def _get_default_format_suggestions(self, data):
        """生成默认的格式建议"""
        if not data or not isinstance(data, list) or not data[0]:
            return None
            
        columns = []
        for header in data[0]:  # 使用第一行作为表头
            columns.append({
                "header": str(header),
                "width": min(len(str(header)) + 4, 40),  # 默认宽度
                "alignment": "center" if header in ["序号", "日期", "时间"] else "left",
                "wrap_text": len(str(header)) > 20,
                "format": "常规",
                "color": "FFE6E6E6"  # 默认表头背景色
            })
        
        return {
            "columns": columns
        }

class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = None
        self.sheet = None
        self._init_workbook()
        
    def _init_workbook(self):
        if os.path.exists(self.excel_path):
            self.workbook = openpyxl.load_workbook(self.excel_path)
        else:
            self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        
    def write_data(self, data, format_suggestions=None):
        """写入数据并应用格式"""
        try:
            if isinstance(data, str):
                parsed_data = json.loads(data)
            else:
                parsed_data = data
                
            if not isinstance(parsed_data, list):
                raise ValueError("数据格式不是数组")
            
            # 清空当前工作表
            self.sheet.delete_rows(1, self.sheet.max_row)
            
            # 写入数据
            for row_idx, row_data in enumerate(parsed_data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = self.sheet.cell(row=row_idx, column=col_idx)
                    cell.value = str(value) if value else ""
            
            # 应用格式建议
            if format_suggestions and 'columns' in format_suggestions:
                for col_idx, col_format in enumerate(format_suggestions['columns'], start=1):
                    column_letter = self.sheet.cell(row=1, column=col_idx).column_letter
                    
                    # 设置列宽
                    if 'width' in col_format:
                        self.sheet.column_dimensions[column_letter].width = col_format['width']
                    
                    # 应用单元格格式
                    for row in self.sheet.iter_rows(min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            # 对齐方式
                            if 'alignment' in col_format:
                                cell.alignment = openpyxl.styles.Alignment(
                                    horizontal=col_format['alignment'],
                                    wrap_text=col_format.get('wrap_text', False)
                                )
                            
                            # 应用颜色
                            if 'color' in col_format:
                                try:
                                    color = col_format['color'].strip().lstrip('#')
                                    # 确保颜色值是8位ARGB格式
                                    if len(color) == 6:
                                        color = 'FF' + color  # 添加不透明度
                                    if cell.row == 1:  # 表头
                                        cell.fill = openpyxl.styles.PatternFill(
                                            start_color=color,
                                            end_color=color,
                                            fill_type='solid'
                                        )
                                except Exception as e:
                                    print(f"设置颜色时出错: {str(e)}")
                            
                            # 设置字体
                            if cell.row == 1:  # 表头加粗
                                cell.font = openpyxl.styles.Font(bold=True)
            else:
                # 默认格式化
                if parsed_data:
                    # 设置默认表头样式
                    for col_idx in range(1, len(parsed_data[0]) + 1):
                        cell = self.sheet.cell(row=1, column=col_idx)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.alignment = openpyxl.styles.Alignment(
                            horizontal='center',
                            vertical='center'
                        )
                        # 设置默认表头背景色
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='FFE6E6E6',  # 浅灰色
                            end_color='FFE6E6E6',
                            fill_type='solid'
                        )
                    
                    # 设置数据行样式
                    for row_idx in range(2, len(parsed_data) + 1):
                        for col_idx in range(1, len(parsed_data[0]) + 1):
                            cell = self.sheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = openpyxl.styles.Alignment(
                                horizontal='left',
                                vertical='center'
                            )
                    
                    # 自动调整列宽
                    for column in self.sheet.columns:
                        max_length = 0
                        column = [cell for cell in column if cell.value]
                        for cell in column:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        if column:
                            adjusted_width = min(max_length + 2, 40)
                            self.sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # 设置表格边框
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            for row in self.sheet.iter_rows(min_row=1, max_row=len(parsed_data)):
                for cell in row:
                    cell.border = thin_border
            
            self.workbook.save(self.excel_path)
            return True
            
        except Exception as e:
            print(f"写入Excel时发生错误: {str(e)}")
            return False

    def write_merged_data(self, data_list):
        try:
            merged_data = []
            for data in data_list:
                if isinstance(data, str):
                    parsed_data = json.loads(data)
                else:
                    parsed_data = data
                
                if not isinstance(parsed_data, list):
                    raise ValueError("数据格式不是数组")
                
                if not merged_data:
                    merged_data.append(parsed_data[0])  # 添加表头
                
                merged_data.extend(parsed_data[1:])  # 添加数据行
            
            # 清空当前工作表
            self.sheet.delete_rows(1, self.sheet.max_row)
            
            # 写入合并后的数据
            for row_idx, row_data in enumerate(merged_data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = self.sheet.cell(row=row_idx, column=col_idx)
                    cell.value = str(value) if value else ""
            
            # 默认格式化
            if merged_data:
                # 设置默认表头样式
                for col_idx in range(1, len(merged_data[0]) + 1):
                    cell = self.sheet.cell(row=1, column=col_idx)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center',
                        vertical='center'
                    )
                    # 设置默认表头背景色
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='FFE6E6E6',  # 浅灰色
                        end_color='FFE6E6E6',
                        fill_type='solid'
                    )
                
                # 设置数据行样式
                for row_idx in range(2, len(merged_data) + 1):
                    for col_idx in range(1, len(merged_data[0]) + 1):
                        cell = self.sheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = openpyxl.styles.Alignment(
                            horizontal='left',
                            vertical='center'
                        )
                
                # 自动调整列宽
                for column in self.sheet.columns:
                    max_length = 0
                    column = [cell for cell in column if cell.value]
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                        if column:  # 确保列不为空
                            adjusted_width = min(max_length + 2, 40)
                            self.sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            self.workbook.save(self.excel_path)
            return True
            
        except Exception as e:
            print(f"合并写入Excel时发生错误: {str(e)}")
            return False

def main():
    try:
        # 从配置文件加载 API Key
        config_path = "config.json"
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                api_key = config["阿里"]["api_key"]
        except Exception:
            print("无法加载配置文件，或配置文件中未找到 API Key")
            return

        if not api_key:
            print("错误：API Key 未设置。请在配置文件中设置有效的 API Key")
            return

        # 初始化图片分析器
        analyzer = ImageAnalyzer(api_key=api_key)
        
        # 测试本地图片
        local_image = r"C:\Users\Cheng-MaoMao\Desktop\PictureTransferForm\微信图片_20250327203511.jpg"
        
        # 根据图片路径生成Excel文件名
        image_name = os.path.splitext(os.path.basename(local_image))[0]
        excel_path = os.path.join(os.path.dirname(local_image), f"{image_name}.xlsx")
        
        # 创建Excel处理器
        excel_handler = ExcelHandler(excel_path)
        
        # 分析图片并处理返回结果
        result = analyzer.analyze_image(local_image)
        if result:
            try:
                # 打印处理前的结果
                print("处理前的结果:", result)
                
                # 移除可能的干扰字符
                result = result.strip().lstrip('\ufeff')
                
                # 确保result不为空且是有效的JSON
                if result:
                    try:
                        # 验证JSON格式
                        json.loads(result)
                        success = excel_handler.write_data(result)
                        if success:
                            print(f"表格已保存至: {excel_path}")
                        else:
                            print("表格保存失败")
                    except json.JSONDecodeError as e:
                        print(f"JSON格式无效: {str(e)}")
                        print("原始数据:", result)
                else:
                    print("分析结果为空")
            except Exception as e:
                print(f"处理结果时出错: {str(e)}")
                print("错误的数据:", result)
        else:
            print("图片分析失败")
    except Exception as e:
        print(f"发生错误: {str(e)}")

if __name__ == "__main__":
    main()